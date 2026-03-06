# reader_databricks.py
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from typing import Dict, Optional, Tuple, List

import re
import pandas as pd
from openpyxl import load_workbook

import config as cfg


@dataclass
class DatabricksContext:
    workbook_path: str

    # SSOT header (01_Advertiser_Name)
    hash_name: str
    tenant_id: str
    account_id: str
    downloaded_dt: Optional[datetime]

    # Export downloaded timestamp anchor (date only)
    ref_date: Optional[date]

    # Derived eval window (SSOT from header Date Range)
    window_start: Optional[date]
    window_end: Optional[date]
    window_days: Optional[int]
    window_str: str

    # Backward-compat: previously used for A1 in Framework_Analysis
    account_name: str

    # sheet_name -> DataFrame
    sheets: Dict[str, pd.DataFrame]


def _norm(s: str) -> str:
    return str(s).strip().lower().replace("\n", " ").replace("\r", " ")


def _parse_datetime_any(x) -> Optional[datetime]:
    if x is None:
        return None
    if isinstance(x, datetime):
        return x
    if isinstance(x, date):
        return datetime(x.year, x.month, x.day)
    s = str(x).strip()
    if not s:
        return None
    dt = pd.to_datetime(s, errors="coerce", utc=False)
    if pd.isna(dt):
        return None
    if isinstance(dt, pd.Timestamp):
        return dt.to_pydatetime()
    return None


def _parse_date_any(x) -> Optional[date]:
    if x is None:
        return None
    if isinstance(x, date) and not isinstance(x, datetime):
        return x
    if isinstance(x, datetime):
        return x.date()
    s = str(x).strip()
    if not s:
        return None
    dt = pd.to_datetime(s, errors="coerce", utc=False)
    if pd.isna(dt):
        return None
    if isinstance(dt, pd.Timestamp):
        return dt.to_pydatetime().date()
    return None


def _extract_header_from_01_advertiser(path: str) -> Tuple[str, str, str, Optional[date], Optional[date], Optional[datetime]]:
    """
    SSOT header extraction from '01_Advertiser_Name' tab.

    Expected patterns in that sheet (based on your exports):
      - A1: "<HashName> - Advertiser_Name"  -> we want "<HashName>"
      - A line containing "Tenant ID:" and "Account ID:"
      - "Date Range:"  -> "YYYY-MM-DD to YYYY-MM-DD"
      - "Downloaded:"  -> "YYYY-MM-DD HH:MM:SS"
    """
    wb = load_workbook(path, read_only=True, data_only=True)

    # Find the 01_Advertiser_Name sheet deterministically
    sheet = None
    for s in wb.sheetnames:
        if str(s).strip().lower().startswith("01_"):
            sheet = s
            break
    if sheet is None:
        # fallback (still deterministic): first sheet that contains both words
        for s in wb.sheetnames:
            sl = str(s).lower()
            if "advertiser" in sl and "name" in sl:
                sheet = s
                break

    if sheet is None:
        wb.close()
        return "", "", "", None, None, None

    ws = wb[sheet]

    # HashName from A1 (strip trailing " - Advertiser_Name")
    a1 = ws["A1"].value
    a1s = str(a1).strip() if a1 is not None else ""
    hash_name = a1s
    if a1s:
        # remove trailing " - Advertiser_Name" (case-insensitive)
        hash_name = re.sub(r"\s-\sAdvertiser_Name\s*$", "", a1s, flags=re.IGNORECASE).strip()

    tenant_id = ""
    account_id = ""
    start = None
    end = None
    downloaded_dt = None

    # scan a small top-left area deterministically
    for r in range(1, 25):
        row_cells = []
        for c in range(1, 15):
            v = ws.cell(r, c).value
            if v is None:
                continue
            row_cells.append(str(v))
        if not row_cells:
            continue

        line = " ".join(row_cells).strip()
        low = line.lower()

        # Tenant/Account IDs (often in the "Account:" line)
        if ("tenant id" in low) and ("account id" in low):
            m_t = re.search(r"Tenant\s*ID:\s*([0-9a-fA-F-]{8,})", line)
            m_a = re.search(r"Account\s*ID:\s*([0-9]{6,})", line)
            if m_t:
                tenant_id = m_t.group(1).strip()
            if m_a:
                account_id = m_a.group(1).strip()

        # Date Range
        if "date range" in low and (start is None or end is None):
            m = re.search(r"(\d{4}-\d{2}-\d{2})\s*to\s*(\d{4}-\d{2}-\d{2})", line, flags=re.IGNORECASE)
            if m:
                start = _parse_date_any(m.group(1))
                end = _parse_date_any(m.group(2))

        # Downloaded timestamp
        if "downloaded" in low and downloaded_dt is None:
            # try full timestamp inside the joined line
            m = re.search(r"(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})", line)
            if m:
                downloaded_dt = _parse_datetime_any(m.group(1))
            else:
                # fallback: if "Downloaded:" is in a cell, take adjacent cell to the right
                for c in range(1, 15):
                    v = ws.cell(r, c).value
                    if v is None:
                        continue
                    if "downloaded" in str(v).lower():
                        adj = ws.cell(r, c + 1).value
                        downloaded_dt = _parse_datetime_any(adj)
                        break

    wb.close()
    return hash_name, tenant_id, account_id, start, end, downloaded_dt


def load_databricks_export(path: str) -> DatabricksContext:
    """
    HARD-LOCKED ingestion rule (per your export standard):
      - Excel row 6 is the header for ALL sheets
      - Excel row 7+ is data
    In pandas, that's header=5 (0-based).
    """
    xls = pd.ExcelFile(path)
    sheets: Dict[str, pd.DataFrame] = {}

    # Load ONLY sheets that match the prefixes listed in cfg.TAB_CANDIDATES (v8 allowlist).
    allowed_prefixes = sorted({p for prefs in cfg.TAB_CANDIDATES.values() for p in prefs})

    def _is_allowed(sheet_name: str) -> bool:
        return any(sheet_name.startswith(p) for p in allowed_prefixes)

    allowed_sheet_names = [s for s in xls.sheet_names if _is_allowed(s)]

    for s in allowed_sheet_names:
        # Hard lock: Excel row 6 is the header (0-based index 5)
        df = pd.read_excel(xls, sheet_name=s, header=5)

        # Drop “Unnamed:” columns caused by merged/title rows above header
        if df is not None and not df.empty:
            df = df.loc[:, ~df.columns.astype(str).str.match(r"^Unnamed:\s*\d+$", case=False)].copy()

        sheets[s] = df

    # ✅ SSOT header extraction (01_Advertiser_Name)
    hash_name, tenant_id, account_id, h_start, h_end, downloaded_dt = _extract_header_from_01_advertiser(path)

    # Anchor ref_date to Downloaded date (per locked agent rules)
    ref_date = downloaded_dt.date() if downloaded_dt else None

    # Window must be copied from header Date Range (SSOT)
    window_start = h_start
    window_end = h_end

    if window_start and window_end:
        window_days = (window_end - window_start).days + 1
        window_str = f"{window_start.isoformat()} to {window_end.isoformat()} ({window_days} days)"
    else:
        window_days = None
        window_str = "UNKNOWN WINDOW (Date Range not found in 01_Advertiser_Name)"

    # Backward-compat: old code expects account_name (used in A1 previously)
    # Now align it to full hash_name so downstream keeps working.
    account_name = hash_name or ""

    return DatabricksContext(
        workbook_path=path,
        hash_name=hash_name,
        tenant_id=tenant_id,
        account_id=account_id,
        downloaded_dt=downloaded_dt,
        ref_date=ref_date,
        account_name=account_name,
        window_start=window_start,
        window_end=window_end,
        window_days=window_days,
        window_str=window_str,
        sheets=sheets,
    )


def get_dataset(ctx: DatabricksContext, dataset_key_or_prefix: str) -> Tuple[Optional[str], Optional[pd.DataFrame]]:
    """
    Returns (sheet_name, df) for a dataset key (preferred) or a tab prefix (fallback).
    Your rules_engine uses this signature.
    """
    # 1) If key exists in TAB_CANDIDATES, search by those prefixes
    cands = cfg.TAB_CANDIDATES.get(dataset_key_or_prefix, [])
    for pref in cands:
        for sname, df in ctx.sheets.items():
            if sname.startswith(pref):
                return sname, df

    # 2) Otherwise treat input as a prefix and attempt match
    pref = dataset_key_or_prefix
    for sname, df in ctx.sheets.items():
        if sname.startswith(pref):
            return sname, df

    return None, None
