# run_agent.py
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from reader_databricks import load_databricks_export
from rules_engine import evaluate_all
from writer_framework import write_results_to_template
from download_helper import create_download_artifact, print_sandbox_link


OUT_DIR = Path("/mnt/data")
MIN_OUTPUT_BYTES = 20000


def _safe_filename(name: str) -> str:
    name = (name or "").strip()
    if not name:
        return "UNKNOWN_ACCOUNT"
    name = re.sub(r'[\\/*?:"<>|]', "", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name or "UNKNOWN_ACCOUNT"


def _read_cell(xlsm_path: str, sheet: str, cell: str) -> str:
    wb = load_workbook(xlsm_path, keep_vba=True, data_only=False)
    try:
        if sheet not in wb.sheetnames:
            return ""
        value = wb[sheet][cell].value
        return str(value).strip() if value is not None else ""
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _assert_file_ok(path: str | Path, min_bytes: int = MIN_OUTPUT_BYTES) -> None:
    p = Path(path)
    if not p.exists():
        raise RuntimeError(f"Output file missing: {p}")
    if p.stat().st_size < min_bytes:
        raise RuntimeError(f"Output file too small: {p} ({p.stat().st_size} bytes)")


def _write_output_metadata(artifact: dict[str, Any], out_dir: Path = OUT_DIR) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    meta_path = out_dir / "framework_output.json"

    payload = {
        "download_path": artifact["download_path"],
        "sandbox_link": artifact["sandbox_link"],
        "download_filename": artifact["download_filename"],
    }

    tmp_path = out_dir / "framework_output.tmp.json"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)
        f.flush()
        os.fsync(f.fileno())

    tmp_path.replace(meta_path)

    if not meta_path.exists():
        raise RuntimeError(f"Failed to create metadata file: {meta_path}")

    return meta_path


def _print_runtime_error(stage: str, reason: str, export_path: str | None, daily_budget: Any) -> None:
    export_name = Path(export_path).name if export_path else "missing"
    budget_value = daily_budget if daily_budget is not None else "missing"

    print("ERROR: Evaluation workflow failed before output generation.")
    print(f"Stage: {stage}")
    print(f"Reason: {reason}")
    print("Inputs received:")
    print(f"- Export file: {export_name}")
    print(f"- Daily budget: {budget_value}")
    print("")
    print("No output file was generated.")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--export",
        required=True,
        help="Databricks export .xlsx (seller_Pre_Analysis_Dashboard*.xlsx)",
    )
    ap.add_argument(
        "--template",
        required=True,
        help="Frozen template .xlsm",
    )
    ap.add_argument(
        "--daily_budget",
        required=True,
        type=float,
        help="Account Daily Budget (number). Required for C050.",
    )
    args = ap.parse_args()

    export_path = args.export
    template_path = args.template
    daily_budget = float(args.daily_budget)

    try:
        if not os.path.exists(export_path):
            raise FileNotFoundError(f"Export not found: {export_path}")
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Template not found: {template_path}")

        OUT_DIR.mkdir(parents=True, exist_ok=True)

        # --------------------------------------------------------------
        # Load export
        # --------------------------------------------------------------
        stage = "reader"
        ctx = load_databricks_export(export_path)
        ctx.daily_budget = daily_budget

        hash_name = getattr(ctx, "hash_name", "") or getattr(ctx, "account_name", "") or "UNKNOWN_ACCOUNT"
        safe_hash = _safe_filename(hash_name)

        # --------------------------------------------------------------
        # Evaluate
        # --------------------------------------------------------------
        stage = "rules_engine"
        results = evaluate_all(ctx)

        # --------------------------------------------------------------
        # Write canonical output
        # --------------------------------------------------------------
        stage = "writer"
        canonical_filename = f"{safe_hash} - Framework Analysis.xlsm"
        canonical_path = OUT_DIR / canonical_filename

        meta = write_results_to_template(
            template_path=template_path,
            output_path=str(canonical_path),
            results=results,
            ctx=ctx,
            sheet_analysis="Framework_Analysis",
            sheet_reference="Framework_Reference",
        )

        # --------------------------------------------------------------
        # Validate canonical output
        # --------------------------------------------------------------
        stage = "save"
        _assert_file_ok(canonical_path, min_bytes=MIN_OUTPUT_BYTES)

        # --------------------------------------------------------------
        # Create fresh timestamped download artifact
        # --------------------------------------------------------------
        stage = "attachment"
        artifact = create_download_artifact(
            canonical_path=str(canonical_path),
            hash_name=safe_hash,
            out_dir=str(OUT_DIR),
            min_bytes=MIN_OUTPUT_BYTES,
        )

        download_path = Path(artifact["download_path"])
        _assert_file_ok(download_path, min_bytes=MIN_OUTPUT_BYTES)

        # --------------------------------------------------------------
        # Metadata file for agent source-of-truth
        # --------------------------------------------------------------
        metadata_path = _write_output_metadata(artifact, out_dir=OUT_DIR)

        # --------------------------------------------------------------
        # Read-back verification
        # --------------------------------------------------------------
        a1 = _read_cell(str(canonical_path), "Framework_Analysis", "A1")
        b3 = _read_cell(str(canonical_path), "Framework_Analysis", "B3")
        b4 = _read_cell(str(canonical_path), "Framework_Analysis", "B4")
        b5 = _read_cell(str(canonical_path), "Framework_Analysis", "B5")

        # --------------------------------------------------------------
        # Structured stdout for the agent
        # --------------------------------------------------------------
        print("DONE")
        print(f"Required output file: {canonical_path}")
        print(f"Download output file: {download_path}")
        print(f"DOWNLOAD_FILENAME: {artifact['download_filename']}")
        print(f"OUTPUT_METADATA: {metadata_path}")
        print(f"Account: {getattr(ctx, 'hash_name', '') or getattr(ctx, 'account_name', '')}")
        print(f"Window: {getattr(ctx, 'window_str', '')}")
        print(f"Ref date: {getattr(ctx, 'ref_date', '')}")
        print(f"Downloaded: {getattr(ctx, 'downloaded_dt', '')}")
        print(f"DailyBudget used: {ctx.daily_budget}")
        print(f"Meta: {meta}")
        print(f"Framework_Analysis!A1: {a1}")
        print(f"Framework_Analysis!B3: {b3}")
        print(f"Framework_Analysis!B4: {b4}")
        print(f"Framework_Analysis!B5: {b5}")

        # Critical lines for reliable GPT download linking
        print_sandbox_link(artifact)

        return 0

    except Exception as e:
        stage_name = locals().get("stage", "unknown")
        _print_runtime_error(
            stage=stage_name,
            reason=str(e),
            export_path=export_path,
            daily_budget=daily_budget,
        )
        return 1


if __name__ == "__main__":
    sys.exit(main())
