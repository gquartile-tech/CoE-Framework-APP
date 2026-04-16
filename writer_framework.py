# writer_framework.py
from __future__ import annotations

from typing import Dict, Tuple, Optional

from openpyxl import load_workbook

OK = "OK"
FLAG = "FLAG"
PARTIAL = "PARTIAL"

# C048 is the only manual control (Cleanup & Relaunch Cadence)
MANUAL_CONTROLS = {"C048"}


def _safe_str(x) -> str:
    return str(x).strip() if x is not None else ""


def _get_hash_name(ctx) -> str:
    if ctx is None:
        return "UNKNOWN ACCOUNT"
    v = _safe_str(getattr(ctx, "hash_name", ""))
    if v:
        return v
    v = _safe_str(getattr(ctx, "account_name", ""))
    return v or "UNKNOWN ACCOUNT"


def _get_tenant_account_line(ctx) -> str:
    if ctx is None:
        return ""
    tenant_id = _safe_str(getattr(ctx, "tenant_id", ""))
    account_id = _safe_str(getattr(ctx, "account_id", ""))
    parts = []
    if tenant_id:
        parts.append(f"Tenant ID: {tenant_id}")
    if account_id:
        parts.append(f"Account ID: {account_id}")
    return " | ".join(parts)


def _get_eval_window_line(ctx) -> str:
    if ctx is None:
        return "UNKNOWN WINDOW"
    v = _safe_str(getattr(ctx, "window_str", ""))
    return v or "UNKNOWN WINDOW"


def _get_downloaded_line(ctx) -> str:
    if ctx is None:
        return ""
    dt = getattr(ctx, "downloaded_dt", None)
    if dt is None:
        return ""
    try:
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return _safe_str(dt)


def _detect_control_id_column(ws) -> Tuple[int, int]:
    """
    Returns (header_row, control_col).
    Tries to find a header cell containing both 'control' and 'id'.
    Falls back to (1, 1).
    """
    for r in range(1, 60):
        for c in range(1, 60):
            val = ws.cell(r, c).value
            if val and ("control" in str(val).lower()) and ("id" in str(val).lower()):
                return r, c
    return 1, 1


def write_results_to_template(
    template_path: str,
    output_path: str,
    results: dict,
    ctx=None,
    sheet_analysis: str = "Framework_Analysis",
    sheet_reference: str = "Framework_Reference",
) -> Dict[str, object]:
    wb = load_workbook(
        template_path,
        keep_vba=True,
        keep_links=True,
        data_only=False,
    )

    analysis_written = False
    rows_written = 0
    header_row = None
    control_col = None

    try:
        # -----------------------------
        # TAB 1 — Framework_Analysis
        # -----------------------------
        if sheet_analysis in wb.sheetnames and ctx is not None:
            ws_a = wb[sheet_analysis]
            ws_a["A1"].value = _get_hash_name(ctx)
            ws_a["B3"].value = _get_tenant_account_line(ctx) or None
            ws_a["B4"].value = _get_eval_window_line(ctx) or None
            ws_a["B5"].value = _get_downloaded_line(ctx) or None
            analysis_written = True

        # -----------------------------
        # TAB 2 — Framework_Reference
        # -----------------------------
        if sheet_reference not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_reference}' not found.")

        ws = wb[sheet_reference]
        header_row, control_col = _detect_control_id_column(ws)

        # Detect column positions for Status, What We Saw, Why It Matters, Notes
        # by reading the header row
        col_status = None
        col_what = None
        col_why = None
        col_notes = None

        for c in range(1, ws.max_column + 1):
            val = _safe_str(ws.cell(header_row, c).value).lower()
            if val == "status":
                col_status = c
            elif "what we saw" in val:
                col_what = c
            elif "why it matters" in val:
                col_why = c
            elif "notes" in val or "intent" in val:
                col_notes = c

        # Fallback to known column positions if header detection fails
        if col_status is None:
            col_status = 4   # column D
        if col_what is None:
            col_what = 8     # column H
        if col_why is None:
            col_why = 9      # column I
        if col_notes is None:
            col_notes = 14   # column N

        # Build ControlID -> row index
        cid_to_row: Dict[str, int] = {}
        blanks_in_a_row = 0
        row = header_row + 1
        max_row = ws.max_row if ws.max_row and ws.max_row > row else row + 500

        while row <= max_row:
            raw = _safe_str(ws.cell(row, control_col).value).upper()
            if not raw:
                blanks_in_a_row += 1
                if blanks_in_a_row >= 25:
                    break
                row += 1
                continue
            blanks_in_a_row = 0
            cid_to_row[raw] = row
            row += 1

        # Write results
        for cid, result in results.items():
            cid_u = _safe_str(cid).upper()
            r = cid_to_row.get(cid_u)
            if not r:
                continue

            # Support both ControlResult dataclass and dict
            status = getattr(result, "status", None) or (result.get("status") if isinstance(result, dict) else None)
            what   = getattr(result, "what",   None) or (result.get("what")   if isinstance(result, dict) else None)
            why    = getattr(result, "why",    None) or (result.get("why")    if isinstance(result, dict) else None)
            # Legacy fallback: old ControlResult used 'note' instead of 'what'
            if not what:
                what = getattr(result, "note", None) or (result.get("note") if isinstance(result, dict) else None)

            status_u = _safe_str(status).upper()
            what_s   = _safe_str(what)
            why_s    = _safe_str(why)

            # Column D — STATUS
            ws.cell(r, col_status).value = status_u if status_u else None

            # Column H — What We Saw (always written)
            ws.cell(r, col_what).value = what_s if what_s else None

            # Column I — Why It Matters (always written)
            ws.cell(r, col_why).value = why_s if why_s else None

            # Column N — Notes/Intent (FLAG/PARTIAL only, manual controls get fixed label)
            if cid_u in MANUAL_CONTROLS:
                final_note = "MANUAL CHECK REQUIRED"
            elif status_u in {FLAG, PARTIAL} and what_s:
                final_note = what_s
            else:
                final_note = ""
            ws.cell(r, col_notes).value = final_note if final_note else None

            rows_written += 1

        wb.save(output_path)

    finally:
        try:
            wb.close()
        except Exception:
            pass

    return {
        "output_path": output_path,
        "analysis_written": analysis_written,
        "reference_rows_written": rows_written,
        "header_row": header_row,
        "control_col": control_col,
    }
